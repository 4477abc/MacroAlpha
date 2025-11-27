#!/usr/bin/env python3
"""
Fix missing dates in Column A of financials Excel files.

Structure (Excel row numbers, 1-based):
- Row 6: Header "Dates" - DO NOT MODIFY
- Row 7: First data row (2005-12-31 for annual, 2005-03-31 for quarterly)
- For ANNUAL: rows 7-26 = years 2005-2024
- For QUARTERLY: rows 7-86 = 80 quarters (2005-Q1 to 2024-Q4)
"""

from openpyxl import load_workbook
from datetime import date

def fix_date_column(filename, period_type):
    print(f"\n{'='*60}")
    print(f"📊 Processing: {filename}")
    print(f"{'='*60}")
    
    wb = load_workbook(filename)
    ws = wb.active
    
    # Show before state
    print("\nBefore (Column A):")
    for row in range(6, 15):
        val = ws.cell(row=row, column=1).value
        print(f"  Row {row}: {val}")
    print("  ...")
    for row in range(20, 28):
        val = ws.cell(row=row, column=1).value
        print(f"  Row {row}: {val}")
    
    # Keep header row intact, fill dates starting from row 7
    if period_type == 'ANNUAL':
        base_year = 2005
        for row in range(7, 28):
            year = base_year + (row - 7)
            if year > 2024:
                break
            ws.cell(row=row, column=1).value = date(year, 12, 31)
            
    elif period_type == 'QUARTERLY':
        quarters = [(3, 31), (6, 30), (9, 30), (12, 31)]
        base_year = 2005
        for row in range(7, 88):
            quarter_idx = row - 7
            year = base_year + (quarter_idx // 4)
            q = quarter_idx % 4
            if year > 2024:
                break
            month, day = quarters[q]
            ws.cell(row=row, column=1).value = date(year, month, day)
    
    wb.save(filename)
    
    # Verify
    wb2 = load_workbook(filename)
    ws2 = wb2.active
    print("\nAfter (Column A):")
    for row in range(6, 15):
        val = ws2.cell(row=row, column=1).value
        if hasattr(val, 'strftime'):
            print(f"  Row {row}: {val.strftime('%Y-%m-%d')}")
        else:
            print(f"  Row {row}: {val}")
    print("  ...")
    
    end_row = 27 if period_type == 'ANNUAL' else 87
    for row in range(end_row - 5, end_row + 1):
        val = ws2.cell(row=row, column=1).value
        if hasattr(val, 'strftime'):
            print(f"  Row {row}: {val.strftime('%Y-%m-%d')}")
        elif val:
            print(f"  Row {row}: {val}")
    
    print(f"\n✅ Saved: {filename}")


if __name__ == "__main__":
    files = [
        ("financials_annual.xlsx", "ANNUAL"),
        ("financials_quarterly.xlsx", "QUARTERLY"),
        ("financials_fiscal_annual.xlsx", "ANNUAL"),
        ("financials_fiscal_quarterly.xlsx", "QUARTERLY"),
    ]
    
    for filename, period_type in files:
        try:
            fix_date_column(filename, period_type)
        except Exception as e:
            print(f"\n❌ Error: {filename}: {e}")
    
    print("\n" + "="*60)
    print("✅ All files processed!")
    print("="*60)

